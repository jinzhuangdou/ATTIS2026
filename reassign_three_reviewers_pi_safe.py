#!/usr/bin/env python3
"""
Assign exactly three reviewers per abstract (cols 5, 10, 15). Clear column 20.
Preserve every cell that already reads "Garmire, Lana" (same row/column unchanged).
Avoid assigning reviewers to abstracts from their own lab PI (Sheet0 col 1 by source_row),
except preserved Lana cells are never modified (per stakeholder request).

Run from review_site/:  python reassign_three_reviewers_pi_safe.py

Uses a fixed RNG seed so the assignment is reproducible; edit SEED to reshuffle.
"""

from __future__ import annotations

import random
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
WORKBOOK = SCRIPT_DIR / "ATTIS abstract submission_April 21, 2026_09.24_with_review_assignments.xlsx"
BACKUP = SCRIPT_DIR / (
    "ATTIS abstract submission_April 21, 2026_09.24_with_review_assignments_BACKUP_before_pi_reassign_three.xlsx"
)

REVIEWER_POOL = [
    "Chen, Jake Y",
    "Garmire, Lana",
    "Chen, Jin (Campus)",
    "Mosa, Abu S",
    "Chong, Zechen",
    "Osborne, John D",
    "Jinzhuang Dou",
    "Amy Wang",
]

LANA = "Garmire, Lana"
REVIEW_COLS = (5, 10, 15)
EXTRA_COL_CLEAR = 20
SEED = 42


def safe(v) -> str:
    return "" if v is None else str(v).strip()


def norm_name(s: str) -> str:
    return " ".join(str(s).strip().lower().split())


def is_lana(val) -> bool:
    if val is None or str(val).strip() == "":
        return False
    return norm_name(val) == norm_name(LANA)


def forbidden_for_pi(pi_raw: str) -> set[str]:
    """Reviewers who must not occupy flexible slots for this PI."""
    p = (pi_raw or "").strip().lower()
    if not p or p == "na":
        return set()
    bad: set[str] = set()

    compact = re.sub(r"\s+", "", p)
    if "lana" in p or "garmire" in p:
        bad.add("Garmire, Lana")
    if re.search(r"\bjake\b", p):
        bad.add("Chen, Jake Y")
    if "zechen" in p or re.search(r"\bchong\b", p):
        bad.add("Chong, Zechen")
    if "jinzhuang" not in compact and re.search(r"\bjin\b", p):
        bad.add("Chen, Jin (Campus)")
    if "jinzhuang" in compact:
        bad.add("Jinzhuang Dou")
    if re.search(r"\bjohn\b", p) or "osborne" in p:
        bad.add("Osborne, John D")
    if "mosa" in p or "abu" in p:
        bad.add("Mosa, Abu S")
    if ("amy" in p and "wang" in p) or re.search(r"\bamywang\b", compact):
        bad.add("Amy Wang")

    return bad


def reassign_row(ra, row: int, pi: str, rng: random.Random) -> None:
    fixed: dict[int, str] = {}
    for col in REVIEW_COLS:
        v = ra.cell(row=row, column=col).value
        if is_lana(v):
            fixed[col] = LANA

    flex_cols = [c for c in REVIEW_COLS if c not in fixed]
    forbid = forbidden_for_pi(pi)
    used = set(fixed.values())
    need = len(flex_cols)

    pool = [r for r in REVIEWER_POOL if r not in used and r != LANA and r not in forbid]
    if len(pool) < need:
        preview = [ra.cell(row=row, column=c).value for c in REVIEW_COLS]
        raise ValueError(
            f"Row {row}: PI={pi!r} need {need} flex reviewers but pool has {len(pool)}. "
            f"forbid={sorted(forbid)}, fixed_Lana_cols={sorted(fixed)}, current={preview}"
        )

    chosen = rng.sample(pool, need)
    for col, name in zip(sorted(flex_cols), chosen):
        ra.cell(row=row, column=col).value = name

    ra.cell(row=row, column=EXTRA_COL_CLEAR).value = None


def main() -> None:
    if not BACKUP.exists():
        shutil.copy2(WORKBOOK, BACKUP)
        print("Created backup:", BACKUP.name)

    rng = random.Random(SEED)

    wb = load_workbook(WORKBOOK, data_only=False)
    s0 = wb["Sheet0"]
    ra = wb["Review_Assignments"]

    def pi_for_source_row(sr) -> str:
        if sr is None:
            return ""
        try:
            sr_int = int(sr)
        except (TypeError, ValueError):
            return ""
        if 2 <= sr_int <= s0.max_row:
            return safe(s0.cell(sr_int, 1).value)
        return ""

    processed = 0
    for row in range(2, ra.max_row + 1):
        if not ra.cell(row=row, column=1).value:
            continue
        sr = ra.cell(row=row, column=2).value
        pi = pi_for_source_row(sr)
        if not pi:
            pi = safe(ra.cell(row=row, column=3).value)

        reassign_row(ra, row, pi, rng)
        processed += 1

    wb.save(WORKBOOK)
    wb.close()
    print(f"Saved {WORKBOOK.name}, rows reassigned: {processed}, seed={SEED}")


if __name__ == "__main__":
    main()
