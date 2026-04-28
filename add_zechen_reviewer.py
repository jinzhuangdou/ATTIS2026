#!/usr/bin/env python3
"""Attach Chong, Zechen as Reviewer 4 (column 20) without changing Lana in cols 5/10/15."""

import shutil
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
SRC = SCRIPT_DIR / "ATTIS abstract submission_April 21, 2026_09.24_with_review_assignments.xlsx"
BACKUP = (
    SCRIPT_DIR
    / "ATTIS abstract submission_April 21, 2026_09.24_with_review_assignments_BACKUP_before_adding_Zechen.xlsx"
)

REV4_COL = 20
ZECHEN = "Chong, Zechen"


def main() -> None:
    shutil.copy2(SRC, BACKUP)
    print("Backup written:", BACKUP.name)

    wb = load_workbook(SRC, data_only=False)
    ra = wb["Review_Assignments"]

    n = 0
    for r in range(2, ra.max_row + 1):
        if not ra.cell(row=r, column=1).value:
            continue
        prev = ra.cell(row=r, column=REV4_COL).value
        if prev and str(prev).strip():
            raise SystemExit(
                f"Row {r} already has Reviewer 4: {prev!r}; stop to avoid overwriting."
            )
        ra.cell(row=r, column=REV4_COL).value = ZECHEN
        n += 1

    wb.save(SRC)
    wb.close()
    print(f"Set {ZECHEN} as Reviewer 4 on {n} abstract rows (cols 5/10/15 unchanged).")


if __name__ == "__main__":
    main()
