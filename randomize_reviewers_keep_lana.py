#!/usr/bin/env python3
"""
Re-randomize reviewer slots (cols 5, 10, 15, 20) for each abstract.
Cells that already contain Garmire, Lana are left unchanged; all other
reviewer cells in that row are filled with a random assignment without
duplicates within the row (sampled from REVIEWER_POOL).
"""

from __future__ import annotations

import random
import shutil
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
WORKBOOK = SCRIPT_DIR / "ATTIS abstract submission_April 21, 2026_09.24_with_review_assignments.xlsx"
BACKUP = (
    SCRIPT_DIR
    / "ATTIS abstract submission_April 21, 2026_09.24_with_review_assignments_BACKUP_before_randomize_keep_lana.xlsx"
)

# Canonical names (must match Excel / portal)
REVIEWER_POOL = [
    "Chen, Jake Y",
    "Garmire, Lana",
    "Chen, Jin (Campus)",
    "Mosa, Abu S",
    "Chong, Zechen",
    "Osborne, John D",
    "Jinzhuang Dou",
    "Amy Wang",
]

LANA = "Garmire, Lana"

REVIEWER_COLS = [5, 10, 15, 20]


def norm_name(s: str) -> str:
    return " ".join(str(s).strip().lower().split())


def is_lana(val) -> bool:
    if val is None or str(val).strip() == "":
        return False
    return norm_name(val) == norm_name(LANA)


def randomize_row(ra, row: int) -> tuple[int, int]:
    """Returns (n_preserved_lana, n_filled)."""
    fixed = {}
    for col in REVIEWER_COLS:
        v = ra.cell(row=row, column=col).value
        if is_lana(v):
            fixed[col] = LANA

    used = set(fixed.values())
    flex_cols = [c for c in REVIEWER_COLS if c not in fixed]
    need = len(flex_cols)

    # Do not assign Lana to flex slots — only preserved cells keep her (same abstracts as before).
    pool = [p for p in REVIEWER_POOL if p != LANA and p not in used]
    if len(pool) < need:
        raise ValueError(
            f"Row {row}: need {need} reviewers but only {len(pool)} available after preserving Lana"
        )

    chosen = random.sample(pool, need)
    for col, name in zip(sorted(flex_cols), chosen):
        ra.cell(row=row, column=col).value = name

    return len(fixed), need


def main() -> None:
    shutil.copy2(WORKBOOK, BACKUP)
    print("Backup:", BACKUP.name)

    wb = load_workbook(WORKBOOK, data_only=False)
    ra = wb["Review_Assignments"]

    total_lana = 0
    total_flex = 0
    for r in range(2, ra.max_row + 1):
        if not ra.cell(row=r, column=1).value:
            continue
        n_lana, n_flex = randomize_row(ra, r)
        total_lana += n_lana
        total_flex += n_flex

    wb.save(WORKBOOK)
    wb.close()
    print(f"Processed rows: Lana cells preserved = {total_lana}, other slots re-randomized = {total_flex}")
    print("Saved:", WORKBOOK.name)


if __name__ == "__main__":
    main()
