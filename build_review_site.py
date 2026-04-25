#!/usr/bin/env python3
import json
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(
    "/Users/doujinzhuang/Library/CloudStorage/OneDrive-UAB-TheUniversityofAlabamaatBirmingham/DouLab_UAB/Service/ATTIS/final"
)
WORKBOOK_PATH = BASE_DIR / "ATTIS abstract submission_April 21, 2026_09.24_with_review_assignments.xlsx"
OUTPUT_HTML = BASE_DIR / "review_site" / "index.html"


def safe(value):
    if value is None:
        return ""
    return str(value).strip()


def load_data():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    src = wb["Sheet0"]
    assign = wb["Review_Assignments"]

    source_rows = {}
    for row in range(2, src.max_row + 1):
        source_rows[row] = {
            "pi": safe(src.cell(row, 1).value),
            "title": safe(src.cell(row, 2).value),
            "authors": safe(src.cell(row, 3).value),
            "affiliations": safe(src.cell(row, 4).value),
            "email": safe(src.cell(row, 5).value),
            "phone": safe(src.cell(row, 6).value),
            "abstract": safe(src.cell(row, 7).value),
            "keywords": safe(src.cell(row, 8).value),
        }

    abstracts = []
    reviewer_set = set()
    for row in range(2, assign.max_row + 1):
        abstract_id = safe(assign.cell(row, 1).value)
        source_row = assign.cell(row, 2).value
        details = source_rows.get(source_row, {})

        reviewers = []
        for base_col in range(5, assign.max_column + 1, 5):
            reviewer = safe(assign.cell(row, base_col).value)
            if reviewer:
                reviewers.append(reviewer)
                reviewer_set.add(reviewer)

        abstracts.append(
            {
                "abstract_id": abstract_id,
                "source_row": source_row,
                "pi": details.get("pi", safe(assign.cell(row, 3).value)),
                "title": details.get("title", safe(assign.cell(row, 4).value)),
                "authors": details.get("authors", ""),
                "affiliations": details.get("affiliations", ""),
                "email": details.get("email", ""),
                "phone": details.get("phone", ""),
                "abstract": details.get("abstract", ""),
                "keywords": details.get("keywords", ""),
                "reviewers": reviewers,
            }
        )

    return sorted(list(reviewer_set)), abstracts


def build_html(reviewers, abstracts):
    reviewers_json = json.dumps(reviewers, ensure_ascii=False)
    abstracts_json = json.dumps(abstracts, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>ATTIS Abstract Review Portal</title>
  <style>
    :root {{
      --bg: #f5f7fb;
      --card: #ffffff;
      --text: #10243e;
      --muted: #5f7189;
      --line: #dce3ef;
      --brand: #1e5ea8;
      --brand2: #0f7f75;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
      background: var(--bg);
      color: var(--text);
    }}
    .wrap {{
      max-width: 1200px;
      margin: 0 auto;
      padding: 20px;
    }}
    .top {{
      background: linear-gradient(135deg, var(--brand), var(--brand2));
      color: #fff;
      border-radius: 14px;
      padding: 18px 20px;
      margin-bottom: 16px;
    }}
    .top h1 {{ margin: 0; font-size: 24px; }}
    .top p {{ margin: 6px 0 0 0; opacity: 0.95; }}
    .toolbar {{
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 14px;
      display: grid;
      gap: 10px;
      grid-template-columns: 260px 1fr auto auto auto;
      align-items: center;
      margin-bottom: 14px;
    }}
    select, input[type="search"], button {{
      width: 100%;
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 9px 10px;
      font-size: 14px;
      background: #fff;
    }}
    button {{
      width: auto;
      cursor: pointer;
      background: #fff;
      font-weight: 600;
    }}
    button.primary {{
      background: var(--brand);
      color: #fff;
      border-color: var(--brand);
    }}
    .meta {{
      font-size: 13px;
      color: var(--muted);
      margin-bottom: 10px;
    }}
    .grid {{
      display: grid;
      gap: 12px;
    }}
    .card {{
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 12px;
      overflow: hidden;
    }}
    .card-head {{
      padding: 12px 14px;
      border-bottom: 1px solid var(--line);
      background: #f9fbff;
    }}
    .title {{
      font-size: 17px;
      font-weight: 700;
      margin-bottom: 5px;
    }}
    .chips {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
    }}
    .chip {{
      font-size: 12px;
      border-radius: 999px;
      background: #eaf1fd;
      color: #184f90;
      padding: 4px 10px;
      border: 1px solid #d3e0f7;
    }}
    .card-body {{
      padding: 12px 14px;
      display: grid;
      grid-template-columns: 1.4fr 1fr;
      gap: 14px;
    }}
    .section h3 {{
      margin: 0 0 8px;
      font-size: 13px;
      letter-spacing: 0.02em;
      color: var(--muted);
      text-transform: uppercase;
    }}
    .text-block {{
      white-space: pre-wrap;
      line-height: 1.45;
      font-size: 14px;
    }}
    .small {{
      font-size: 13px;
      color: var(--muted);
      margin: 2px 0;
      white-space: pre-wrap;
    }}
    .scores {{
      display: grid;
      grid-template-columns: 1fr 100px;
      gap: 8px 10px;
      align-items: center;
    }}
    .scores label {{
      font-size: 14px;
      font-weight: 600;
    }}
    .scores input {{
      text-align: center;
    }}
    .total {{
      margin-top: 8px;
      font-weight: 700;
      color: #174f8d;
    }}
    .note {{
      width: 100%;
      min-height: 72px;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 8px;
      font-family: inherit;
      resize: vertical;
    }}
    .empty {{
      background: #fff;
      border: 1px dashed var(--line);
      border-radius: 12px;
      padding: 20px;
      text-align: center;
      color: var(--muted);
    }}
    @media (max-width: 980px) {{
      .toolbar {{ grid-template-columns: 1fr; }}
      .card-body {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="top">
      <h1>ATTIS Abstract Review Portal</h1>
      <p>Select reviewer, read full abstract, score 4 criteria (0-10), and export your scores.</p>
    </div>

    <div class="toolbar">
      <select id="reviewerSelect"></select>
      <input id="searchBox" type="search" placeholder="Search by title, PI, keyword..." />
      <button id="resetBtn">Reset Current Reviewer Scores</button>
      <button id="exportReviewerBtn" class="primary">Export Reviewer CSV</button>
      <button id="exportAllBtn">Export All CSV</button>
    </div>

    <div class="meta" id="metaLine"></div>
    <div id="list" class="grid"></div>
  </div>

  <script>
    const REVIEWERS = {reviewers_json};
    const ABSTRACTS = {abstracts_json};
    const STORAGE_KEY = "attis_review_scores_v1";

    const reviewerSelect = document.getElementById("reviewerSelect");
    const searchBox = document.getElementById("searchBox");
    const list = document.getElementById("list");
    const metaLine = document.getElementById("metaLine");
    const resetBtn = document.getElementById("resetBtn");
    const exportReviewerBtn = document.getElementById("exportReviewerBtn");
    const exportAllBtn = document.getElementById("exportAllBtn");

    function getState() {{
      try {{
        return JSON.parse(localStorage.getItem(STORAGE_KEY) || "{{}}");
      }} catch (_) {{
        return {{}};
      }}
    }}

    function setState(next) {{
      localStorage.setItem(STORAGE_KEY, JSON.stringify(next));
    }}

    function stateKey(reviewer, abstractId) {{
      return `${{reviewer}}::${{abstractId}}`;
    }}

    function esc(value) {{
      return (value || "").replace(/[&<>"]/g, (ch) => ({{ "&": "&amp;", "<": "&lt;", ">": "&gt;", '"': "&quot;" }}[ch]));
    }}

    function criterionInput(abstractId, reviewer, field, value) {{
      return `<input type="number" min="0" max="10" step="1" data-aid="${{abstractId}}" data-reviewer="${{esc(reviewer)}}" data-field="${{field}}" value="${{value ?? ""}}" />`;
    }}

    function criteriaTotal(entry) {{
      const n = (x) => Number.isFinite(Number(x)) ? Number(x) : 0;
      return n(entry.topic_fitness) + n(entry.approach) + n(entry.results) + n(entry.innovation);
    }}

    function render() {{
      const reviewer = reviewerSelect.value;
      const query = searchBox.value.trim().toLowerCase();
      const state = getState();

      const filtered = ABSTRACTS.filter((a) => {{
        if (!a.reviewers.includes(reviewer)) return false;
        if (!query) return true;
        const blob = [a.title, a.pi, a.keywords, a.authors, a.abstract].join(" ").toLowerCase();
        return blob.includes(query);
      }});

      metaLine.textContent = `Reviewer: ${{reviewer}} | Assigned abstracts: ${{filtered.length}}`;

      if (!filtered.length) {{
        list.innerHTML = `<div class="empty">No assigned abstracts matched your current search.</div>`;
        return;
      }}

      list.innerHTML = filtered.map((a) => {{
        const key = stateKey(reviewer, a.abstract_id);
        const entry = state[key] || {{}};
        const total = criteriaTotal(entry);
        return `
          <article class="card">
            <div class="card-head">
              <div class="title">${{esc(a.title || "(No title)")}}</div>
              <div class="chips">
                <span class="chip">Abstract #${{esc(a.abstract_id)}}</span>
                <span class="chip">PI: ${{esc(a.pi || "N/A")}}</span>
                <span class="chip">Keywords: ${{esc(a.keywords || "N/A")}}</span>
              </div>
            </div>
            <div class="card-body">
              <section class="section">
                <h3>Abstract</h3>
                <div class="text-block">${{esc(a.abstract || "N/A")}}</div>
                <h3 style="margin-top:12px;">Authors</h3>
                <div class="small">${{esc(a.authors || "N/A")}}</div>
                <h3 style="margin-top:12px;">Affiliations</h3>
                <div class="small">${{esc(a.affiliations || "N/A")}}</div>
              </section>
              <section class="section">
                <h3>Scoring (0-10 each)</h3>
                <div class="scores">
                  <label>Topic fitness</label>${{criterionInput(a.abstract_id, reviewer, "topic_fitness", entry.topic_fitness)}}
                  <label>Approach</label>${{criterionInput(a.abstract_id, reviewer, "approach", entry.approach)}}
                  <label>Results</label>${{criterionInput(a.abstract_id, reviewer, "results", entry.results)}}
                  <label>Innovation</label>${{criterionInput(a.abstract_id, reviewer, "innovation", entry.innovation)}}
                </div>
                <div class="total">Total: <span data-total="${{a.abstract_id}}">${{total}}</span> / 40</div>
                <h3 style="margin-top:12px;">Optional Notes</h3>
                <textarea class="note" data-note-aid="${{a.abstract_id}}" data-reviewer="${{esc(reviewer)}}" placeholder="Short reviewer note...">${{esc(entry.note || "")}}</textarea>
              </section>
            </div>
          </article>
        `;
      }}).join("");
    }}

    function clampScore(value) {{
      const n = Number(value);
      if (!Number.isFinite(n)) return "";
      if (n < 0) return 0;
      if (n > 10) return 10;
      return Math.round(n);
    }}

    function updateEntry(abstractId, reviewer, field, value) {{
      const state = getState();
      const key = stateKey(reviewer, abstractId);
      const entry = state[key] || {{}};
      entry[field] = value;
      state[key] = entry;
      setState(state);
    }}

    function csvEscape(val) {{
      const s = String(val ?? "");
      if (s.includes('"') || s.includes(",") || s.includes("\\n")) {{
        return `"${{s.replace(/"/g, '""')}}"`;
      }}
      return s;
    }}

    function downloadCsv(filename, rows) {{
      const content = rows.map((r) => r.map(csvEscape).join(",")).join("\\n");
      const blob = new Blob([content], {{ type: "text/csv;charset=utf-8;" }});
      const url = URL.createObjectURL(blob);
      const a = document.createElement("a");
      a.href = url;
      a.download = filename;
      document.body.appendChild(a);
      a.click();
      a.remove();
      URL.revokeObjectURL(url);
    }}

    function exportRows(targetReviewer = null) {{
      const state = getState();
      const rows = [[
        "Reviewer",
        "Abstract ID",
        "PI",
        "Title",
        "Topic fitness",
        "Approach",
        "Results",
        "Innovation",
        "Total",
        "Note"
      ]];

      for (const a of ABSTRACTS) {{
        const reviewers = targetReviewer ? [targetReviewer] : a.reviewers;
        for (const reviewer of reviewers) {{
          if (!a.reviewers.includes(reviewer)) continue;
          const entry = state[stateKey(reviewer, a.abstract_id)] || {{}};
          const total = criteriaTotal(entry);
          rows.push([
            reviewer,
            a.abstract_id,
            a.pi,
            a.title,
            entry.topic_fitness ?? "",
            entry.approach ?? "",
            entry.results ?? "",
            entry.innovation ?? "",
            total,
            entry.note ?? ""
          ]);
        }}
      }}
      return rows;
    }}

    reviewerSelect.innerHTML = REVIEWERS.map((r) => `<option value="${{esc(r)}}">${{esc(r)}}</option>`).join("");

    reviewerSelect.addEventListener("change", render);
    searchBox.addEventListener("input", render);

    list.addEventListener("input", (e) => {{
      const t = e.target;
      if (t.matches("input[type='number'][data-aid][data-reviewer][data-field]")) {{
        const v = clampScore(t.value);
        t.value = v;
        updateEntry(t.dataset.aid, t.dataset.reviewer, t.dataset.field, v);
        const state = getState();
        const entry = state[stateKey(t.dataset.reviewer, t.dataset.aid)] || {{}};
        const totalEl = document.querySelector(`[data-total="${{t.dataset.aid}}"]`);
        if (totalEl) totalEl.textContent = criteriaTotal(entry);
      }}
      if (t.matches("textarea[data-note-aid][data-reviewer]")) {{
        updateEntry(t.dataset.noteAid, t.dataset.reviewer, "note", t.value);
      }}
    }});

    resetBtn.addEventListener("click", () => {{
      const reviewer = reviewerSelect.value;
      if (!confirm(`Reset all saved scores for ${{reviewer}}?`)) return;
      const state = getState();
      for (const a of ABSTRACTS) {{
        delete state[stateKey(reviewer, a.abstract_id)];
      }}
      setState(state);
      render();
    }});

    exportReviewerBtn.addEventListener("click", () => {{
      const reviewer = reviewerSelect.value;
      const rows = exportRows(reviewer);
      const safeName = reviewer.replace(/[^a-z0-9]+/gi, "_");
      downloadCsv(`ATTIS_review_scores_${{safeName}}.csv`, rows);
    }});

    exportAllBtn.addEventListener("click", () => {{
      const rows = exportRows(null);
      downloadCsv("ATTIS_review_scores_all.csv", rows);
    }});

    render();
  </script>
</body>
</html>
"""


def main():
    reviewers, abstracts = load_data()
    OUTPUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    html = build_html(reviewers, abstracts)
    OUTPUT_HTML.write_text(html, encoding="utf-8")
    print(f"Built {OUTPUT_HTML}")
    print(f"Reviewers: {len(reviewers)}")
    print(f"Abstracts: {len(abstracts)}")


if __name__ == "__main__":
    main()
