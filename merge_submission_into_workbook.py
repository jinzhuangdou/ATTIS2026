#!/usr/bin/env python3
"""Merge plain submission workbook (Sheet0) into the assignments workbook."""

from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent

SUBMISSION_PATH = SCRIPT_DIR.parent / "ATTIS abstract submission_April 21, 2026_09.24.xlsx"
ASSIGNMENTS_PATH = SCRIPT_DIR / "ATTIS abstract submission_April 21, 2026_09.24_with_review_assignments.xlsx"

# Provisional reviewers for newly appended abstracts (cols 5, 10, 15). Edit in Excel if needed.
DEFAULT_NEW_REVIEWERS = ("Garmire, Lana", "Osborne, John D", "Chen, Jake Y")


def titled_sheet_rows(ws) -> list[int]:
    return [
        r
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=2).value and str(ws.cell(row=r, column=2).value).strip()
    ]


def count_ra_rows(assign_sheet) -> int:
    return sum(
        1
        for r in range(2, assign_sheet.max_row + 1)
        if assign_sheet.cell(row=r, column=1).value not in (None, "")
    )


def main() -> None:
    sub = load_workbook(SUBMISSION_PATH, data_only=False)
    src_sub = sub.active
    wb = load_workbook(ASSIGNMENTS_PATH, data_only=False)
    s0_ass = wb["Sheet0"]
    ra = wb["Review_Assignments"]

    titled = titled_sheet_rows(src_sub)
    for row in range(1, src_sub.max_row + 1):
        for col in range(1, 9):
            s0_ass.cell(row=row, column=col).value = src_sub.cell(row=row, column=col).value

    n_ra = count_ra_rows(ra)
    need = len(titled) - n_ra
    print(f"Titled abstracts: {len(titled)}")
    print(f"Review_Assignments rows: {n_ra}")

    if need > 0:
        print(f"Appending {need} provisional RA row(s).")
        for k in range(need):
            last_r = max(
                r for r in range(2, ra.max_row + 1) if ra.cell(row=r, column=1).value not in (None, "")
            )
            new_r = last_r + 1
            idx = n_ra + k
            source_row_num = titled[idx]
            aid = idx + 1
            pi = src_sub.cell(row=source_row_num, column=1).value or ""
            title = src_sub.cell(row=source_row_num, column=2).value or ""
            ra.cell(row=new_r, column=1).value = str(aid)
            ra.cell(row=new_r, column=2).value = source_row_num
            ra.cell(row=new_r, column=3).value = str(pi).strip()
            ra.cell(row=new_r, column=4).value = str(title).strip()
            ra.cell(row=new_r, column=5).value = DEFAULT_NEW_REVIEWERS[0]
            ra.cell(row=new_r, column=10).value = DEFAULT_NEW_REVIEWERS[1]
            ra.cell(row=new_r, column=15).value = DEFAULT_NEW_REVIEWERS[2]
            print(
                f"  Row {new_r}: id={aid} source_row={source_row_num} "
                f"reviewers={DEFAULT_NEW_REVIEWERS}"
            )
    wb.save(ASSIGNMENTS_PATH)
    print("Saved:", ASSIGNMENTS_PATH)


if __name__ == "__main__":
    main()
