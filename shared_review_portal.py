#!/usr/bin/env python3
import csv
import io
import json
import hashlib
import os
import secrets
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

from fastapi import FastAPI, Header, HTTPException, Query
from fastapi.responses import HTMLResponse, PlainTextResponse, StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field


DEFAULT_BASE_DIR = Path(__file__).resolve().parent
BASE_DIR = Path(os.getenv("ATTIS_BASE_DIR", str(DEFAULT_BASE_DIR)))
WORKBOOK_PATH = Path(
    os.getenv(
        "ATTIS_WORKBOOK_PATH",
        str(BASE_DIR / "ATTIS abstract submission_April 21, 2026_09.24_with_review_assignments.xlsx"),
    )
)
DB_PATH = Path(os.getenv("ATTIS_DB_PATH", str(BASE_DIR / "review_portal.db")))
PIN_EXPORT_PATH = Path(os.getenv("ATTIS_PIN_EXPORT_PATH", str(BASE_DIR / "reviewer_pins.csv")))

# Default reviewer passwords (Blazer ID format).
# These are applied on startup so deployed instances stay consistent.
DEFAULT_REVIEWER_PASSWORDS = {
    "Chen, Jake Y": "jakechen",
    "Garmire, Lana": "lgarmire",
    "Chen, Jin (Campus)": "jinchen",
    "Mosa, Abu S": "asmosa",
    "Chong, Zechen": "zchong",
    "Osborne, John D": "josborne",
    "Jinzhuang Dou": "jdou",
    "Amy Wang": "amywang",
}


def get_reviewer_password_map() -> Dict[str, str]:
    raw = os.getenv("REVIEWER_PASSWORDS_JSON", "").strip()
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                return {str(k): str(v) for k, v in parsed.items()}
        except Exception as e:
            print(f"[ATTIS] Invalid REVIEWER_PASSWORDS_JSON, fallback to defaults: {e}")
    return DEFAULT_REVIEWER_PASSWORDS


def safe(value) -> str:
    return "" if value is None else str(value).strip()


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    conn.execute("PRAGMA busy_timeout = 30000;")
    return conn


def init_db() -> None:
    with get_conn() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS abstracts (
              abstract_id TEXT PRIMARY KEY,
              source_row INTEGER,
              pi TEXT,
              title TEXT,
              authors TEXT,
              affiliations TEXT,
              email TEXT,
              phone TEXT,
              abstract_body TEXT,
              keywords TEXT
            );

            CREATE TABLE IF NOT EXISTS assignments (
              abstract_id TEXT NOT NULL,
              reviewer TEXT NOT NULL,
              PRIMARY KEY (abstract_id, reviewer),
              FOREIGN KEY (abstract_id) REFERENCES abstracts(abstract_id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS reviews (
              abstract_id TEXT NOT NULL,
              reviewer TEXT NOT NULL,
              topic_fitness INTEGER,
              approach INTEGER,
              results INTEGER,
              innovation INTEGER,
              note TEXT,
              updated_at TEXT NOT NULL,
              PRIMARY KEY (abstract_id, reviewer),
              FOREIGN KEY (abstract_id, reviewer) REFERENCES assignments(abstract_id, reviewer) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS reviewer_auth (
              reviewer TEXT PRIMARY KEY,
              pin_hash TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS sessions (
              token TEXT PRIMARY KEY,
              reviewer TEXT NOT NULL,
              expires_at TEXT NOT NULL
            );
            """
        )


def _hash_pin(pin: str) -> str:
    return hashlib.sha256(pin.encode("utf-8")).hexdigest()


def _ensure_reviewer_auth(conn: sqlite3.Connection) -> List[Dict[str, str]]:
    reviewers = [
        r["reviewer"]
        for r in conn.execute(
            "SELECT DISTINCT reviewer FROM assignments ORDER BY reviewer COLLATE NOCASE"
        ).fetchall()
    ]
    password_map = get_reviewer_password_map()
    pin_rows: List[Dict[str, str]] = []
    for reviewer in reviewers:
        pin = password_map.get(reviewer)
        if not pin:
            # Fallback for any future reviewer not in the fixed list.
            pin = "".join(secrets.choice("0123456789") for _ in range(6))
        conn.execute(
            "INSERT OR REPLACE INTO reviewer_auth (reviewer, pin_hash) VALUES (?, ?)",
            (reviewer, _hash_pin(pin)),
        )
        # Invalidate old sessions so password updates are effective immediately.
        conn.execute("DELETE FROM sessions WHERE reviewer = ?", (reviewer,))
        pin_rows.append({"reviewer": reviewer, "pin": pin})

    placeholders = ",".join("?" for _ in reviewers) or "''"
    conn.execute(
        f"DELETE FROM reviewer_auth WHERE reviewer NOT IN ({placeholders})",
        reviewers,
    )
    conn.execute(
        f"DELETE FROM sessions WHERE reviewer NOT IN ({placeholders})",
        reviewers,
    )
    return pin_rows


def _write_pin_file(pin_rows: List[Dict[str, str]]) -> None:
    if not pin_rows:
        return
    PIN_EXPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with PIN_EXPORT_PATH.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Reviewer", "PIN", "Note"])
        for row in pin_rows:
            note = (
                "New PIN generated; share with reviewer"
                if row["pin"] != "UNCHANGED"
                else "Already exists from previous run"
            )
            writer.writerow([row["reviewer"], row["pin"], note])


def sync_from_workbook() -> Dict[str, int]:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    src = wb["Sheet0"]
    assign = wb["Review_Assignments"]

    source_rows = {}
    for row in range(2, src.max_row + 1):
        source_rows[row] = {
            "pi": safe(src.cell(row, 1).value),
            "title": safe(src.cell(row, 2).value),
            "authors": safe(src.cell(row, 3).value),
            "affiliations": safe(src.cell(row, 4).value),
            "email": safe(src.cell(row, 5).value),
            "phone": safe(src.cell(row, 6).value),
            "abstract_body": safe(src.cell(row, 7).value),
            "keywords": safe(src.cell(row, 8).value),
        }

    abstracts = []
    assignments = []
    for row in range(2, assign.max_row + 1):
        abstract_id = safe(assign.cell(row, 1).value)
        if not abstract_id:
            continue

        source_row = assign.cell(row, 2).value
        details = source_rows.get(source_row, {})
        abstract = {
            "abstract_id": abstract_id,
            "source_row": source_row,
            "pi": details.get("pi", safe(assign.cell(row, 3).value)),
            "title": details.get("title", safe(assign.cell(row, 4).value)),
            "authors": details.get("authors", ""),
            "affiliations": details.get("affiliations", ""),
            "email": details.get("email", ""),
            "phone": details.get("phone", ""),
            "abstract_body": details.get("abstract_body", ""),
            "keywords": details.get("keywords", ""),
        }
        abstracts.append(abstract)

        for base_col in range(5, assign.max_column + 1, 5):
            reviewer = safe(assign.cell(row, base_col).value)
            if reviewer:
                assignments.append((abstract_id, reviewer))

    with get_conn() as conn:
        conn.execute("DELETE FROM assignments")
        conn.execute("DELETE FROM abstracts")
        conn.executemany(
            """
            INSERT INTO abstracts (
              abstract_id, source_row, pi, title, authors, affiliations,
              email, phone, abstract_body, keywords
            ) VALUES (
              :abstract_id, :source_row, :pi, :title, :authors, :affiliations,
              :email, :phone, :abstract_body, :keywords
            )
            """,
            abstracts,
        )
        conn.executemany(
            "INSERT INTO assignments (abstract_id, reviewer) VALUES (?, ?)",
            assignments,
        )
        conn.execute(
            """
            DELETE FROM reviews
            WHERE (abstract_id, reviewer) NOT IN (
              SELECT abstract_id, reviewer FROM assignments
            )
            """
        )
        pin_rows = _ensure_reviewer_auth(conn)
    _write_pin_file(pin_rows)
    return {"abstracts": len(abstracts), "assignments": len(assignments)}


class ReviewUpsert(BaseModel):
    abstract_id: str
    topic_fitness: Optional[int] = Field(default=None, ge=0, le=10)
    approach: Optional[int] = Field(default=None, ge=0, le=10)
    results: Optional[int] = Field(default=None, ge=0, le=10)
    innovation: Optional[int] = Field(default=None, ge=0, le=10)
    note: Optional[str] = ""


class LoginRequest(BaseModel):
    reviewer: str
    pin: str


def _reviewer_from_token(token: Optional[str]) -> str:
    if not token:
        raise HTTPException(status_code=401, detail="Missing auth token")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT reviewer
            FROM sessions
            WHERE token = ? AND expires_at >= ?
            """,
            (token, now),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    return row["reviewer"]


app = FastAPI(title="ATTIS 2026 abstract shared review panel")


@app.on_event("startup")
def startup() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    init_db()
    if WORKBOOK_PATH.exists():
        sync_from_workbook()
    else:
        print(f"[ATTIS] Workbook not found at startup: {WORKBOOK_PATH}")


@app.get("/", response_class=HTMLResponse)
def index() -> str:
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>ATTIS 2026 abstract shared review panel</title>
  <style>
    :root { --bg:#f4f7fc; --card:#fff; --line:#d7e0ef; --brand:#1f5fa8; --text:#10243e; --muted:#5b6f88; --ok:#1d7c52; }
    * { box-sizing:border-box; }
    body { margin:0; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Arial,sans-serif; background:var(--bg); color:var(--text); }
    .wrap { max-width:1200px; margin:0 auto; padding:18px; }
    .hero { background:linear-gradient(135deg,#1f5fa8,#0f7f75); color:#fff; border-radius:14px; padding:16px 18px; margin-bottom:14px; }
    .hero h1 { margin:0; font-size:24px; }
    .hero p { margin:6px 0 0; opacity:.95; }
    .toolbar { display:grid; grid-template-columns:280px 1fr auto auto auto; gap:10px; background:var(--card); border:1px solid var(--line); border-radius:12px; padding:12px; margin-bottom:10px; }
    select,input,button { border:1px solid var(--line); border-radius:10px; padding:9px 10px; font-size:14px; background:#fff; }
    button { font-weight:600; cursor:pointer; }
    button.primary { background:var(--brand); color:#fff; border-color:var(--brand); }
    .tabs { display:flex; gap:10px; margin:10px 0; }
    .tab { border:1px solid var(--line); background:#fff; border-radius:999px; padding:8px 12px; cursor:pointer; font-weight:700; color:#32547b; }
    .tab.active { background:#e8f0fd; border-color:#b8cff3; color:#1f4f90; }
    .meta { font-size:13px; color:var(--muted); margin:6px 0 10px; }
    .grid { display:grid; gap:12px; }
    .card { background:var(--card); border:1px solid var(--line); border-radius:12px; overflow:hidden; }
    .head { padding:12px 14px; border-bottom:1px solid var(--line); background:#f8fbff; }
    .title { font-size:17px; font-weight:800; margin-bottom:6px; }
    .chips { display:flex; flex-wrap:wrap; gap:8px; }
    .chip { font-size:12px; padding:4px 10px; border-radius:999px; border:1px solid #d2dff4; background:#eaf1fd; color:#194f8f; }
    .body { padding:12px 14px; display:grid; grid-template-columns:1.45fr 1fr; gap:12px; }
    .label { font-size:12px; text-transform:uppercase; letter-spacing:.03em; color:var(--muted); margin-bottom:6px; font-weight:700; }
    .text { white-space:pre-wrap; line-height:1.45; font-size:14px; }
    .scores { display:grid; grid-template-columns:1fr 110px; gap:8px; align-items:center; }
    .scores input { text-align:center; }
    .note { width:100%; min-height:70px; border:1px solid var(--line); border-radius:8px; padding:8px; resize:vertical; font-family:inherit; }
    .save-row { display:flex; align-items:center; gap:10px; margin-top:10px; }
    .save-state { font-size:12px; color:var(--ok); font-weight:700; }
    .summary-card { background:#fff; border:1px solid var(--line); border-radius:12px; padding:12px; overflow:auto; }
    table { border-collapse:collapse; width:100%; font-size:13px; }
    th,td { border:1px solid var(--line); padding:7px 8px; text-align:left; }
    th { background:#f2f7ff; }
    .hidden { display:none; }
    @media (max-width:980px) { .toolbar { grid-template-columns:1fr; } .body { grid-template-columns:1fr; } }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="hero">
      <h1>ATTIS 2026 abstract shared review panel</h1>
      <p>Shared link, server-side save, editable scores, and live overall summary.</p>
    </div>

    <div class="toolbar" style="grid-template-columns: 320px 140px auto auto 1fr;">
      <select id="loginReviewerSelect"></select>
      <input id="pinInput" type="password" placeholder="Password (Blazer ID)" />
      <button id="loginBtn" class="primary">Login</button>
      <button id="logoutBtn">Logout</button>
      <div class="meta" id="authStatus" style="margin:0;"></div>
    </div>

    <div class="tabs">
      <button class="tab active" id="reviewTab">Reviewer View</button>
      <button class="tab" id="adminTab">Overall Summary</button>
    </div>

    <section id="reviewSection">
      <div class="toolbar" style="grid-template-columns: 1fr auto auto;">
        <input id="searchBox" placeholder="Search title, PI, keywords..." />
        <button id="reloadBtn">Reload</button>
        <button id="exportBtn" class="primary">Export Reviewer CSV</button>
      </div>
      <div class="meta" id="metaLine"></div>
      <div id="cards" class="grid"></div>
    </section>

    <section id="adminSection" class="hidden">
      <div class="toolbar" style="grid-template-columns:auto;">
        <button id="refreshSummaryBtn" class="primary">Refresh Summary</button>
      </div>
      <div class="summary-card">
        <div class="label">Per-abstract summary</div>
        <div id="summaryTable"></div>
      </div>
      <div class="summary-card" style="margin-top:12px;">
        <div class="label">Reviewer progress</div>
        <div id="progressTable"></div>
      </div>
    </section>
  </div>

  <script>
    const loginReviewerSelect = document.getElementById("loginReviewerSelect");
    const pinInput = document.getElementById("pinInput");
    const loginBtn = document.getElementById("loginBtn");
    const logoutBtn = document.getElementById("logoutBtn");
    const authStatus = document.getElementById("authStatus");
    const searchBox = document.getElementById("searchBox");
    const cards = document.getElementById("cards");
    const metaLine = document.getElementById("metaLine");
    const reviewTab = document.getElementById("reviewTab");
    const adminTab = document.getElementById("adminTab");
    const reviewSection = document.getElementById("reviewSection");
    const adminSection = document.getElementById("adminSection");
    const summaryTable = document.getElementById("summaryTable");
    const progressTable = document.getElementById("progressTable");

    let reviewers = [];
    let current = [];
    let authToken = localStorage.getItem("attis_auth_token") || "";
    let currentReviewer = localStorage.getItem("attis_reviewer") || "";

    function esc(s) {
      return String(s ?? "").replace(/[&<>"]/g, (c) => ({ "&":"&amp;", "<":"&lt;", ">":"&gt;", '"':"&quot;" }[c]));
    }

    function total(r) {
      const n = (x) => Number.isFinite(Number(x)) ? Number(x) : 0;
      return n(r.topic_fitness)+n(r.approach)+n(r.results)+n(r.innovation);
    }

    function reviewerFromUrl() {
      const p = new URLSearchParams(location.search);
      return p.get("reviewer");
    }

    function authHeaders(extra = {}) {
      if (!authToken) return extra;
      return { ...extra, "X-Auth-Token": authToken };
    }

    async function loadReviewers() {
      const res = await fetch("/api/reviewers");
      const data = await res.json();
      reviewers = data.reviewers || [];
      loginReviewerSelect.innerHTML = reviewers.map(r => `<option value="${esc(r)}">${esc(r)}</option>`).join("");
      const rUrl = reviewerFromUrl();
      if (rUrl && reviewers.includes(rUrl) && !currentReviewer) loginReviewerSelect.value = rUrl;
      if (currentReviewer && reviewers.includes(currentReviewer)) loginReviewerSelect.value = currentReviewer;
    }

    function renderAuthState() {
      if (authToken && currentReviewer) {
        authStatus.textContent = `Logged in as: ${currentReviewer}`;
        loginReviewerSelect.value = currentReviewer;
        loginReviewerSelect.disabled = true;
        pinInput.disabled = true;
        loginBtn.disabled = true;
        logoutBtn.disabled = false;
      } else {
        authStatus.textContent = "Not logged in";
        loginReviewerSelect.disabled = false;
        pinInput.disabled = false;
        loginBtn.disabled = false;
        logoutBtn.disabled = true;
      }
    }

    async function login() {
      const reviewer = loginReviewerSelect.value;
      const pin = pinInput.value.trim();
      if (!reviewer || !pin) {
        alert("Please select reviewer and enter password.");
        return;
      }
      const res = await fetch("/api/login", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ reviewer, pin })
      });
      if (!res.ok) {
        alert("Login failed: invalid reviewer or password.");
        return;
      }
      const data = await res.json();
      authToken = data.token;
      currentReviewer = data.reviewer;
      localStorage.setItem("attis_auth_token", authToken);
      localStorage.setItem("attis_reviewer", currentReviewer);
      pinInput.value = "";
      renderAuthState();
      await loadAssigned();
    }

    async function logout() {
      if (authToken) {
        await fetch("/api/logout", { method: "POST", headers: authHeaders() });
      }
      authToken = "";
      currentReviewer = "";
      localStorage.removeItem("attis_auth_token");
      localStorage.removeItem("attis_reviewer");
      current = [];
      cards.innerHTML = `<div class="card"><div class="head">Please login to view your assigned abstracts.</div></div>`;
      metaLine.textContent = "";
      renderAuthState();
    }

    async function loadAssigned() {
      if (!authToken) {
        cards.innerHTML = `<div class="card"><div class="head">Please login to view your assigned abstracts.</div></div>`;
        metaLine.textContent = "";
        return;
      }
      const q = encodeURIComponent(searchBox.value.trim());
      const res = await fetch(`/api/assigned?q=${q}`, { headers: authHeaders() });
      if (res.status === 401) {
        await logout();
        alert("Session expired. Please login again.");
        return;
      }
      const data = await res.json();
      current = data.items || [];
      metaLine.textContent = `Reviewer: ${currentReviewer} | Assigned abstracts: ${current.length}`;
      renderCards();
    }

    function numberInput(v, id, field) {
      return `<input type="number" min="0" max="10" step="1" value="${v ?? ""}" data-id="${id}" data-field="${field}" />`;
    }

    function renderCards() {
      if (!current.length) {
        cards.innerHTML = `<div class="card"><div class="head">No abstracts found for this reviewer/filter.</div></div>`;
        return;
      }
      cards.innerHTML = current.map(a => `
        <article class="card">
          <div class="head">
            <div class="title">${esc(a.title || "(No title)")}</div>
            <div class="chips">
              <span class="chip">Abstract #${esc(a.abstract_id)}</span>
              <span class="chip">PI: ${esc(a.pi || "N/A")}</span>
              <span class="chip">Keywords: ${esc(a.keywords || "N/A")}</span>
            </div>
          </div>
          <div class="body">
            <section>
              <div class="label">Abstract</div>
              <div class="text">${esc(a.abstract_body || "N/A")}</div>
              <div class="label" style="margin-top:10px;">Authors</div>
              <div class="text">${esc(a.authors || "N/A")}</div>
              <div class="label" style="margin-top:10px;">Affiliations</div>
              <div class="text">${esc(a.affiliations || "N/A")}</div>
            </section>
            <section>
              <div class="label">Scoring (0-10 each)</div>
              <div class="scores">
                <label>Topic fitness</label>${numberInput(a.topic_fitness, a.abstract_id, "topic_fitness")}
                <label>Approach</label>${numberInput(a.approach, a.abstract_id, "approach")}
                <label>Results</label>${numberInput(a.results, a.abstract_id, "results")}
                <label>Innovation</label>${numberInput(a.innovation, a.abstract_id, "innovation")}
              </div>
              <div style="margin-top:8px;font-weight:700;color:#1f4f90;">Total: <span data-total="${a.abstract_id}">${total(a)}</span> / 40</div>
              <div class="label" style="margin-top:10px;">Optional Notes</div>
              <textarea class="note" data-note-id="${a.abstract_id}">${esc(a.note || "")}</textarea>
              <div class="save-row">
                <button class="primary" data-save-id="${a.abstract_id}">Save</button>
                <span class="save-state" id="state-${a.abstract_id}">${a.updated_at ? `Last saved: ${esc(a.updated_at)}` : ""}</span>
              </div>
            </section>
          </div>
        </article>
      `).join("");
    }

    function getCardPayload(abstractId) {
      const card = [...document.querySelectorAll(`input[data-id="${abstractId}"]`)];
      const note = document.querySelector(`textarea[data-note-id="${abstractId}"]`);
      const payload = { abstract_id: abstractId, note: note ? note.value : "" };
      for (const input of card) {
        const v = input.value === "" ? null : Math.max(0, Math.min(10, Math.round(Number(input.value))));
        input.value = v === null ? "" : v;
        payload[input.dataset.field] = v;
      }
      return payload;
    }

    async function saveReview(abstractId) {
      if (!authToken) {
        alert("Please login first.");
        return;
      }
      const payload = getCardPayload(abstractId);
      const res = await fetch("/api/reviews", {
        method: "PUT",
        headers: authHeaders({ "Content-Type": "application/json" }),
        body: JSON.stringify(payload)
      });
      if (res.status === 401) {
        await logout();
        alert("Session expired. Please login again.");
        return;
      }
      if (!res.ok) {
        alert("Save failed.");
        return;
      }
      const data = await res.json();
      const state = document.getElementById(`state-${abstractId}`);
      if (state) state.textContent = `Last saved: ${data.updated_at}`;
      const item = current.find(x => x.abstract_id === abstractId);
      if (item) {
        item.topic_fitness = payload.topic_fitness;
        item.approach = payload.approach;
        item.results = payload.results;
        item.innovation = payload.innovation;
        const totalEl = document.querySelector(`[data-total="${abstractId}"]`);
        if (totalEl) totalEl.textContent = total(item);
      }
    }

    async function loadSummary() {
      const res = await fetch("/api/summary");
      const data = await res.json();

      summaryTable.innerHTML = `
        <table>
          <thead><tr><th>Abstract ID</th><th>PI</th><th>Title</th><th>Assigned Reviewers (#)</th><th>Assigned Reviewer Names</th><th>Submitted Reviews</th><th>Avg Total (/40)</th><th>Avg Topic</th><th>Avg Approach</th><th>Avg Results</th><th>Avg Innovation</th></tr></thead>
          <tbody>
            ${data.abstracts.map(a => `<tr>
              <td>${esc(a.abstract_id)}</td>
              <td>${esc(a.pi)}</td>
              <td>${esc(a.title)}</td>
              <td>${a.assigned_reviewers}</td>
              <td>${esc(a.assigned_reviewer_names || "")}</td>
              <td>${a.review_count}</td>
              <td>${a.avg_total}</td>
              <td>${a.avg_topic_fitness}</td>
              <td>${a.avg_approach}</td>
              <td>${a.avg_results}</td>
              <td>${a.avg_innovation}</td>
            </tr>`).join("")}
          </tbody>
        </table>
      `;

      progressTable.innerHTML = `
        <table>
          <thead><tr><th>Reviewer</th><th>Assigned</th><th>Scored</th><th>Completion</th></tr></thead>
          <tbody>
            ${data.progress.map(p => `<tr>
              <td>${esc(p.reviewer)}</td>
              <td>${p.assigned_count}</td>
              <td>${p.scored_count}</td>
              <td>${p.completion_pct}%</td>
            </tr>`).join("")}
          </tbody>
        </table>
      `;
    }

    reviewTab.addEventListener("click", () => {
      reviewTab.classList.add("active"); adminTab.classList.remove("active");
      reviewSection.classList.remove("hidden"); adminSection.classList.add("hidden");
    });
    adminTab.addEventListener("click", async () => {
      adminTab.classList.add("active"); reviewTab.classList.remove("active");
      adminSection.classList.remove("hidden"); reviewSection.classList.add("hidden");
      await loadSummary();
    });

    loginBtn.addEventListener("click", login);
    logoutBtn.addEventListener("click", logout);
    searchBox.addEventListener("input", loadAssigned);
    document.getElementById("reloadBtn").addEventListener("click", loadAssigned);
    document.getElementById("refreshSummaryBtn").addEventListener("click", loadSummary);

    document.getElementById("exportBtn").addEventListener("click", async () => {
      if (!authToken) {
        alert("Please login first.");
        return;
      }
      const res = await fetch("/api/export.csv", { headers: authHeaders() });
      if (!res.ok) {
        alert("Export failed.");
        return;
      }
      const blob = await res.blob();
      const url = URL.createObjectURL(blob);
      const a = document.createElement("a");
      a.href = url;
      a.download = `ATTIS_scores_${currentReviewer.replace(/[^a-z0-9]+/gi, "_")}.csv`;
      document.body.appendChild(a);
      a.click();
      a.remove();
      URL.revokeObjectURL(url);
    });

    cards.addEventListener("click", async (e) => {
      const btn = e.target.closest("button[data-save-id]");
      if (!btn) return;
      await saveReview(btn.dataset.saveId);
    });

    cards.addEventListener("input", (e) => {
      const input = e.target;
      if (input.matches("input[type='number'][data-id][data-field]")) {
        const abstractId = input.dataset.id;
        const vals = [...document.querySelectorAll(`input[data-id="${abstractId}"]`)].map(x => Number(x.value) || 0);
        const t = vals.reduce((a,b)=>a+b,0);
        const totalEl = document.querySelector(`[data-total="${abstractId}"]`);
        if (totalEl) totalEl.textContent = t;
      }
    });

    (async function init() {
      await loadReviewers();
      renderAuthState();
      if (authToken && currentReviewer) {
        await loadAssigned();
      } else {
        cards.innerHTML = `<div class="card"><div class="head">Please login to view your assigned abstracts.</div></div>`;
      }
      const urlReviewer = reviewerFromUrl();
      if (urlReviewer && !currentReviewer) {
        const lock = document.createElement("div");
        lock.className = "meta";
        lock.textContent = `Reviewer pre-selected by link: ${urlReviewer}`;
        document.querySelector(".hero").appendChild(lock);
      }
    })();
  </script>
</body>
</html>"""


@app.get("/api/reviewers")
def get_reviewers():
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT DISTINCT reviewer FROM assignments ORDER BY reviewer COLLATE NOCASE"
        ).fetchall()
    return {"reviewers": [r["reviewer"] for r in rows]}


@app.post("/api/login")
def login(payload: LoginRequest):
    with get_conn() as conn:
        row = conn.execute(
            "SELECT pin_hash FROM reviewer_auth WHERE reviewer = ?",
            (payload.reviewer,),
        ).fetchone()
        if not row or row["pin_hash"] != _hash_pin(payload.pin.strip()):
            raise HTTPException(status_code=401, detail="Invalid reviewer or PIN")
        token = secrets.token_urlsafe(32)
        expires_at = (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d %H:%M:%S")
        conn.execute(
            "INSERT OR REPLACE INTO sessions (token, reviewer, expires_at) VALUES (?, ?, ?)",
            (token, payload.reviewer, expires_at),
        )
    return {"token": token, "reviewer": payload.reviewer, "expires_at": expires_at}


@app.post("/api/logout")
def logout(x_auth_token: Optional[str] = Header(default=None, alias="X-Auth-Token")):
    if x_auth_token:
        with get_conn() as conn:
            conn.execute("DELETE FROM sessions WHERE token = ?", (x_auth_token,))
    return {"ok": True}


@app.get("/api/assigned")
def get_assigned(
    q: str = Query(default=""),
    x_auth_token: Optional[str] = Header(default=None, alias="X-Auth-Token"),
):
    reviewer = _reviewer_from_token(x_auth_token)
    query = """
    SELECT
      a.abstract_id, a.source_row, a.pi, a.title, a.authors, a.affiliations,
      a.email, a.phone, a.abstract_body, a.keywords,
      r.topic_fitness, r.approach, r.results, r.innovation, r.note, r.updated_at
    FROM assignments x
    JOIN abstracts a ON a.abstract_id = x.abstract_id
    LEFT JOIN reviews r ON r.abstract_id = x.abstract_id AND r.reviewer = x.reviewer
    WHERE x.reviewer = :reviewer
      AND (
        :q = ''
        OR lower(a.title) LIKE '%' || lower(:q) || '%'
        OR lower(a.pi) LIKE '%' || lower(:q) || '%'
        OR lower(a.keywords) LIKE '%' || lower(:q) || '%'
        OR lower(a.abstract_body) LIKE '%' || lower(:q) || '%'
      )
    ORDER BY CAST(a.abstract_id AS INTEGER)
    """
    with get_conn() as conn:
        rows = conn.execute(query, {"reviewer": reviewer, "q": q}).fetchall()
    return {"items": [dict(r) for r in rows]}


@app.put("/api/reviews")
def upsert_review(
    payload: ReviewUpsert,
    x_auth_token: Optional[str] = Header(default=None, alias="X-Auth-Token"),
):
    reviewer = _reviewer_from_token(x_auth_token)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_conn() as conn:
        exists = conn.execute(
            """
            SELECT 1 FROM assignments
            WHERE abstract_id = ? AND reviewer = ?
            """,
            (payload.abstract_id, reviewer),
        ).fetchone()
        if not exists:
            raise HTTPException(status_code=400, detail="Reviewer is not assigned to this abstract")

        conn.execute(
            """
            INSERT INTO reviews (
              abstract_id, reviewer, topic_fitness, approach, results, innovation, note, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(abstract_id, reviewer) DO UPDATE SET
              topic_fitness = excluded.topic_fitness,
              approach = excluded.approach,
              results = excluded.results,
              innovation = excluded.innovation,
              note = excluded.note,
              updated_at = excluded.updated_at
            """,
            (
                payload.abstract_id,
                reviewer,
                payload.topic_fitness,
                payload.approach,
                payload.results,
                payload.innovation,
                payload.note or "",
                now,
            ),
        )
    return {"ok": True, "updated_at": now}


@app.get("/api/summary")
def summary():
    with get_conn() as conn:
        abstract_rows = conn.execute(
            """
            WITH assigned_counts AS (
              SELECT abstract_id, COUNT(*) AS assigned_reviewers
              FROM assignments
              GROUP BY abstract_id
            ),
            assigned_names AS (
              SELECT abstract_id, GROUP_CONCAT(reviewer, '; ') AS assigned_reviewer_names
              FROM (
                SELECT abstract_id, reviewer
                FROM assignments
                ORDER BY abstract_id, reviewer COLLATE NOCASE
              )
              GROUP BY abstract_id
            ),
            review_stats AS (
              SELECT
                abstract_id,
                COUNT(*) AS review_count,
                ROUND(AVG(COALESCE(topic_fitness, 0)), 2) AS avg_topic_fitness,
                ROUND(AVG(COALESCE(approach, 0)), 2) AS avg_approach,
                ROUND(AVG(COALESCE(results, 0)), 2) AS avg_results,
                ROUND(AVG(COALESCE(innovation, 0)), 2) AS avg_innovation,
                ROUND(AVG(COALESCE(topic_fitness, 0) + COALESCE(approach, 0) + COALESCE(results, 0) + COALESCE(innovation, 0)), 2) AS avg_total
              FROM reviews
              GROUP BY abstract_id
            )
            SELECT
              a.abstract_id, a.pi, a.title,
              COALESCE(ac.assigned_reviewers, 0) AS assigned_reviewers,
              COALESCE(an.assigned_reviewer_names, '') AS assigned_reviewer_names,
              COALESCE(rs.review_count, 0) AS review_count,
              COALESCE(rs.avg_topic_fitness, 0) AS avg_topic_fitness,
              COALESCE(rs.avg_approach, 0) AS avg_approach,
              COALESCE(rs.avg_results, 0) AS avg_results,
              COALESCE(rs.avg_innovation, 0) AS avg_innovation,
              COALESCE(rs.avg_total, 0) AS avg_total
            FROM abstracts a
            LEFT JOIN assigned_counts ac ON ac.abstract_id = a.abstract_id
            LEFT JOIN assigned_names an ON an.abstract_id = a.abstract_id
            LEFT JOIN review_stats rs ON rs.abstract_id = a.abstract_id
            ORDER BY CAST(a.abstract_id AS INTEGER)
            """
        ).fetchall()

        progress_rows = conn.execute(
            """
            SELECT
              x.reviewer,
              COUNT(*) AS assigned_count,
              SUM(
                CASE WHEN r.topic_fitness IS NOT NULL
                  OR r.approach IS NOT NULL
                  OR r.results IS NOT NULL
                  OR r.innovation IS NOT NULL
                THEN 1 ELSE 0 END
              ) AS scored_count
            FROM assignments x
            LEFT JOIN reviews r
              ON r.abstract_id = x.abstract_id AND r.reviewer = x.reviewer
            GROUP BY x.reviewer
            ORDER BY x.reviewer COLLATE NOCASE
            """
        ).fetchall()

    progress = []
    for r in progress_rows:
        assigned = int(r["assigned_count"] or 0)
        scored = int(r["scored_count"] or 0)
        pct = round((scored / assigned) * 100, 1) if assigned else 0.0
        progress.append(
            {
                "reviewer": r["reviewer"],
                "assigned_count": assigned,
                "scored_count": scored,
                "completion_pct": pct,
            }
        )
    return {"abstracts": [dict(r) for r in abstract_rows], "progress": progress}


@app.post("/api/sync")
def sync():
    try:
        counts = sync_from_workbook()
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True, **counts}


@app.get("/api/export.csv")
def export_csv(x_auth_token: Optional[str] = Header(default=None, alias="X-Auth-Token")):
    reviewer = _reviewer_from_token(x_auth_token)
    query = """
    SELECT
      x.reviewer, a.abstract_id, a.pi, a.title,
      r.topic_fitness, r.approach, r.results, r.innovation,
      (COALESCE(r.topic_fitness,0) + COALESCE(r.approach,0) + COALESCE(r.results,0) + COALESCE(r.innovation,0)) AS total,
      r.note, r.updated_at
    FROM assignments x
    JOIN abstracts a ON a.abstract_id = x.abstract_id
    LEFT JOIN reviews r ON r.abstract_id = x.abstract_id AND r.reviewer = x.reviewer
    """
    params: List[str] = [reviewer]
    query += " WHERE x.reviewer = ? ORDER BY x.reviewer COLLATE NOCASE, CAST(a.abstract_id AS INTEGER)"

    with get_conn() as conn:
        rows = conn.execute(query, params).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "Reviewer",
            "Abstract ID",
            "PI",
            "Title",
            "Topic fitness",
            "Approach",
            "Results",
            "Innovation",
            "Total",
            "Note",
            "Updated at",
        ]
    )
    for r in rows:
        writer.writerow(
            [
                r["reviewer"],
                r["abstract_id"],
                r["pi"],
                r["title"],
                r["topic_fitness"],
                r["approach"],
                r["results"],
                r["innovation"],
                r["total"],
                r["note"],
                r["updated_at"],
            ]
        )
    output.seek(0)
    filename = f"ATTIS_scores_{reviewer.replace(' ', '_')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/health", response_class=PlainTextResponse)
def health():
    return "ok"
