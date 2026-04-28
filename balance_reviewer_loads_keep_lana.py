#!/usr/bin/env python3
"""
Reassign flexible reviewer slots (cols 5,10,15 excluding preserved Garmire,Lana)
so each non-Lana reviewer’s load stays in **[LOAD_MIN, LOAD_MAX]** per run (defaults
10–12). Totals stay consistent with preserved Lana (flex slot count summed over rows).

Preserves: every existing Garmire, Lana cell (row/column/value), PI conflict rules
from `reassign_three_reviewers_pi_safe`, clear column 20.

The optimizer rejects counts outside the band first, then spreads load using random
integer targets that sum exactly to the flex-slot total — **not forced to eleven each**.

Run:  python balance_reviewer_loads_keep_lana.py
"""

from __future__ import annotations

import itertools
import random
import shutil
from collections import Counter
from pathlib import Path
from typing import Dict, List, NamedTuple, Optional, Sequence, Tuple

from openpyxl import load_workbook

from reassign_three_reviewers_pi_safe import (
    EXTRA_COL_CLEAR,
    REVIEW_COLS,
    REVIEWER_POOL,
    WORKBOOK,
    forbidden_for_pi,
    is_lana,
    safe,
)


def norm_name(s: str) -> str:
    return " ".join(str(s).strip().lower().split())

LANA = "Garmire, Lana"
NON_LANA: List[str] = [r for r in REVIEWER_POOL if r != LANA]

# Flexible-slot load bounds for non-Lana reviewers (inclusive).
LOAD_MIN = 10
LOAD_MAX = 12

BACKUP = WORKBOOK.parent / (
    "ATTIS abstract submission_April 21, 2026_09.24_with_review_assignments_BACKUP_before_balance_10_to_12.xlsx"
)


def enumerate_load_patterns(n_reviewers: int, flex_total: int) -> List[Tuple[int, ...]]:
    """All integer tuples of length n_reviewers summing to flex_total with values in [LOAD_MIN, LOAD_MAX]."""
    rng = range(LOAD_MIN, LOAD_MAX + 1)
    out = [tup for tup in itertools.product(rng, repeat=n_reviewers) if sum(tup) == flex_total]
    if not out:
        raise ValueError(f"No feasible load pattern: {flex_total=} {LOAD_MIN}-{LOAD_MAX} for {n_reviewers} reviewers")
    return out


def pick_targets(rng: random.Random, patterns: Sequence[Tuple[int, ...]]) -> Dict[str, int]:
    """Random multiset paired with reviewer names; prefer patterns that are not all eleven."""
    varied = [p for p in patterns if any(v != 11 for v in p)]
    use = rng.choice(varied if varied else list(patterns))
    vals = list(use)
    rng.shuffle(vals)
    return dict(zip(NON_LANA, vals))


def band_violation(cnt: Counter[str]) -> int:
    s = 0
    for r in NON_LANA:
        x = cnt[r]
        if x < LOAD_MIN:
            s += (LOAD_MIN - x) ** 2
        elif x > LOAD_MAX:
            s += (x - LOAD_MAX) ** 2
    return s


def sq_err(cnt: Counter[str], target: Dict[str, int]) -> int:
    return sum((cnt[r] - target[r]) ** 2 for r in NON_LANA)


class RowSpec(NamedTuple):
    excel_row: int
    flex_cols: Tuple[int, ...]
    allowed: frozenset[str]
    pi: str


def load_row_specs(wb) -> List[RowSpec]:
    s0 = wb["Sheet0"]
    ra = wb["Review_Assignments"]

    def pi_for_source_row(sr) -> str:
        if sr is None:
            return ""
        try:
            sr_int = int(sr)
        except (TypeError, ValueError):
            return ""
        if 2 <= sr_int <= s0.max_row:
            return safe(s0.cell(sr_int, 1).value)
        return ""

    specs: List[RowSpec] = []
    for row in range(2, ra.max_row + 1):
        if not ra.cell(row=row, column=1).value:
            continue
        sr = ra.cell(row=row, column=2).value
        pi = pi_for_source_row(sr)
        if not pi:
            pi = safe(ra.cell(row=row, column=3).value)

        flex_cols = tuple(c for c in REVIEW_COLS if not is_lana(ra.cell(row=row, column=c).value))
        forbid = forbidden_for_pi(pi)
        pool = frozenset(r for r in REVIEWER_POOL if r != LANA and r not in forbid)
        k = len(flex_cols)
        if len(pool) < k:
            raise ValueError(f"Row {row}: need {k} reviewers but PI allows only {sorted(pool)}")

        specs.append(RowSpec(row, flex_cols, pool, pi))

    return specs


def greedy_assign(
    specs: Sequence[RowSpec],
    rng: random.Random,
    target: Dict[str, int],
) -> Dict[Tuple[int, int], str]:
    """Assign each flex cell; pick per-row combinations that best match target loads and [LOAD_MIN,LOAD_MAX]."""
    order = list(specs)
    rng.shuffle(order)
    counts: Counter[str] = Counter()
    out: Dict[Tuple[int, int], str] = {}

    for spec in order:
        cols = spec.flex_cols
        k = len(cols)
        if k == 0:
            continue

        allowed = [r for r in NON_LANA if r in spec.allowed]
        if len(allowed) < k:
            raise RuntimeError(f"Row {spec.excel_row}: allowed pool too small")

        best_combo: Optional[Tuple[str, ...]] = None
        best_key: Optional[Tuple[int, int]] = None

        for combo in itertools.combinations(allowed, k):
            tmp = Counter(counts)
            for person in combo:
                tmp[person] += 1
            key = (band_violation(tmp), sq_err(tmp, target))
            if best_key is None or key < best_key:
                best_key = key
                best_combo = combo

        assert best_combo is not None
        for person in best_combo:
            counts[person] += 1

        for col, name in zip(sorted(cols), sorted(best_combo)):
            out[(spec.excel_row, col)] = name

    return out


def all_counts(assign: Dict[Tuple[int, int], str]) -> Counter[str]:
    cnt: Counter[str] = Counter(assign.values())
    return cnt


def lex_score(cnt: Counter[str], target: Dict[str, int]) -> Tuple[int, int, int]:
    """Lower is better: band violations, squared error vs target, spread."""
    vals = [cnt[r] for r in NON_LANA]
    spread = max(vals) - min(vals) if vals else 0
    return band_violation(cnt), sq_err(cnt, target), spread


def feasible_replace(
    spec: RowSpec, assign: Dict[Tuple[int, int], str], col_change: int, new_rev: str
) -> bool:
    others = []
    for c in spec.flex_cols:
        if c == col_change:
            continue
        others.append(assign[(spec.excel_row, c)])
    if new_rev in others:
        return False
    if new_rev not in spec.allowed:
        return False
    return True


def local_search(
    specs: Sequence[RowSpec],
    assign: Dict[Tuple[int, int], str],
    rng: random.Random,
    target: Dict[str, int],
    max_iters: int = 18_000,
) -> Dict[Tuple[int, int], str]:
    assign = dict(assign)
    spec_by_row = {s.excel_row: s for s in specs}

    def score_now() -> Tuple[int, int, int]:
        return lex_score(all_counts(assign), target)

    best_scr = score_now()
    best_map = dict(assign)

    def improve_from(candidate: Dict[Tuple[int, int], str], scr: Tuple[int, int, int]) -> None:
        nonlocal best_scr, best_map
        if scr <= best_scr:
            best_scr = scr
            best_map = dict(candidate)

    for _ in range(max_iters):
        if best_scr[0] == 0 and best_scr[1] == 0:
            break
        if rng.random() < 0.72:
            spec = rng.choice(specs)
            row = spec.excel_row
            col = rng.choice(spec.flex_cols)
            cur = assign[(row, col)]
            alts = list(NON_LANA)
            rng.shuffle(alts)
            for alt in alts:
                if alt == cur:
                    continue
                if not feasible_replace(spec, assign, col, alt):
                    continue
                assign[(row, col)] = alt
                sc = score_now()
                if sc <= best_scr:
                    improve_from(assign, sc)
                else:
                    assign[(row, col)] = cur
        else:
            a, b = rng.sample(specs, 2)
            if not a.flex_cols or not b.flex_cols:
                continue
            ca, cb = rng.choice(a.flex_cols), rng.choice(b.flex_cols)
            ra_name, rb_name = assign[(a.excel_row, ca)], assign[(b.excel_row, cb)]
            if rb_name not in a.allowed or ra_name not in b.allowed:
                continue
            others_a = [assign[(a.excel_row, c)] for c in a.flex_cols if c != ca]
            others_b = [assign[(b.excel_row, c)] for c in b.flex_cols if c != cb]
            if rb_name in others_a or ra_name in others_b:
                continue
            if rb_name == ra_name:
                continue

            assign[(a.excel_row, ca)] = rb_name
            assign[(b.excel_row, cb)] = ra_name
            sc = score_now()
            if sc <= best_scr:
                improve_from(assign, sc)
            else:
                assign[(a.excel_row, ca)] = ra_name
                assign[(b.excel_row, cb)] = rb_name

    return best_map


def verify(
    specs: Sequence[RowSpec],
    assign: Dict[Tuple[int, int], str],
    lana_snap: Dict[Tuple[int, int], str],
) -> None:
    for spec in specs:
        names = []
        for c in REVIEW_COLS:
            snap = (spec.excel_row, c)
            if snap in lana_snap:
                if norm_name(lana_snap[snap]) != norm_name(LANA):
                    raise AssertionError("Lana snapshot mismatch", snap, lana_snap[snap])
                names.append(LANA)
            else:
                names.append(assign[(spec.excel_row, c)])
        if len(set(names)) != 3:
            raise AssertionError(f"Row {spec.excel_row}: duplicate reviewer {names}")
        for n in names:
            if n == LANA:
                continue
            if n not in spec.allowed:
                raise AssertionError(f"Row {spec.excel_row}: {n} not allowed for PI {spec.pi!r}")


def snapshot_lana_cells(ra) -> Dict[Tuple[int, int], str]:
    snap: Dict[Tuple[int, int], str] = {}
    for row in range(2, ra.max_row + 1):
        for col in REVIEW_COLS:
            v = ra.cell(row=row, column=col).value
            if v is None or safe(v) == "":
                continue
            if norm_name(str(v)) == norm_name(LANA):
                snap[(row, col)] = safe(v)
    return snap


def apply_assignment(wb, assign: Dict[Tuple[int, int], str], lana_snap: Dict[Tuple[int, int], str]) -> None:
    ra = wb["Review_Assignments"]
    rows_done = sorted({row for row, _ in assign.keys()} | {r for r, _ in lana_snap.keys()})

    for row in rows_done:
        for col in REVIEW_COLS:
            key = (row, col)
            if key in lana_snap:
                ra.cell(row=row, column=col).value = LANA
            elif key in assign:
                ra.cell(row=row, column=col).value = assign[key]

        ra.cell(row=row, column=EXTRA_COL_CLEAR).value = None


def main() -> None:
    if not BACKUP.exists():
        shutil.copy2(WORKBOOK, BACKUP)
        print("Backup:", BACKUP.name)

    wb = load_workbook(WORKBOOK, data_only=False)
    ra_before = wb["Review_Assignments"]
    lana_snap = snapshot_lana_cells(ra_before)
    specs = load_row_specs(wb)

    flex_total = sum(len(s.flex_cols) for s in specs)
    patterns = enumerate_load_patterns(len(NON_LANA), flex_total)

    best_assign: Dict[Tuple[int, int], str] | None = None
    best_global: Tuple[int, int, int] = (999, 999, 999)

    for restart in range(72):
        rng = random.Random(10_000 + restart)
        target = pick_targets(rng, patterns)
        a0 = greedy_assign(specs, rng, target)
        rng2 = random.Random(restart + 50_000)
        a1 = local_search(specs, a0, rng2, target)
        cnt = all_counts(a1)
        scr = lex_score(cnt, target)
        if scr <= best_global:
            best_global = scr
            best_assign = a1

    assert best_assign is not None
    verify(specs, best_assign, lana_snap)

    cnt_chk = Counter(best_assign.values())
    if band_violation(cnt_chk) != 0:
        raise RuntimeError(f"Could not satisfy loads in [{LOAD_MIN},{LOAD_MAX}]: {dict(cnt_chk)}")

    apply_assignment(wb, best_assign, lana_snap)

    wb.save(WORKBOOK)
    wb.close()

    cnt_final = Counter(best_assign.values())
    print(f"Loads must stay in [{LOAD_MIN}, {LOAD_MAX}], flex slots = {flex_total}")
    print("Objective (violations, sq_err, spread):", best_global)
    print("Flexible-slot loads (non-Lana):")
    for name in NON_LANA:
        print(f"  {name:24} {cnt_final[name]}")
    print("Garmire, Lana (fixed):", len(lana_snap), "abstracts")


if __name__ == "__main__":
    main()
